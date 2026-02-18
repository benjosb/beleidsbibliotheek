#!/usr/bin/env python3
"""
Converteert de Beleidsbibliotheek.xlsx naar JSON-bestanden voor de webinterface.
Leest zowel de verrijkte "Totaal geschoond" data als de jaarspecifieke tabbladen.
"""

import json
import os
from datetime import datetime
import openpyxl

EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', 'Beleidsbibliotheek.xlsx')
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), '..', 'data')

os.makedirs(OUTPUT_DIR, exist_ok=True)

def date_to_str(val):
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    return str(val) if val else None

def clean(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None

def read_enriched_sheet(wb):
    """Leest het 'Totaal geschoond' tabblad met verrijkte metadata."""
    ws = wb['Totaal geschoond']
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        datum, naam, besluit, domein, type_doc, onderwerp, link, juridisch = row[:8]
        if not naam:
            continue
        records.append({
            'datum': date_to_str(datum),
            'naam': clean(naam),
            'besluit': clean(besluit),
            'domein': clean(domein),
            'type_document': clean(type_doc),
            'onderwerp_begroting': clean(onderwerp),
            'link': clean(link),
            'juridische_classificatie': clean(juridisch),
            'bron': 'raad',
            'type_besluit': 'Raadsbesluit'
        })
    return records

def read_year_sheet(wb, year):
    """Leest een jaartabblad."""
    sheet_name = str(year)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    records = []
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    has_onderwerp = 'Onderwerp begroting' in (headers or [])

    for row in ws.iter_rows(min_row=2, values_only=True):
        if has_onderwerp and len(row) >= 5:
            datum, naam, besluit, onderwerp, link = row[0], row[1], row[2], row[3], row[4]
            status = None
        elif len(row) >= 5:
            datum, naam, besluit, status, link = row[0], row[1], row[2], row[3], row[4]
            onderwerp = None
        else:
            continue

        if not naam:
            continue

        record = {
            'datum': date_to_str(datum),
            'naam': clean(naam),
            'besluit': clean(besluit),
            'link': clean(link),
            'bron': 'raad',
            'type_besluit': 'Raadsbesluit',
            'jaar': year
        }
        if status:
            record['status'] = clean(status)
        if onderwerp:
            record['onderwerp_begroting'] = clean(onderwerp)
        records.append(record)
    return records

def enrich_year_data(year_records, enriched_records):
    """Verrijkt jaar-records met metadata uit 'Totaal geschoond' op basis van naam-matching."""
    enriched_by_name = {}
    for r in enriched_records:
        if r['naam']:
            enriched_by_name[r['naam'].lower().strip()] = r

    for record in year_records:
        key = record['naam'].lower().strip() if record['naam'] else ''
        if key in enriched_by_name:
            source = enriched_by_name[key]
            for field in ['domein', 'type_document', 'onderwerp_begroting', 'juridische_classificatie']:
                if field not in record or not record.get(field):
                    if source.get(field):
                        record[field] = source[field]
    return year_records

def build_thema_boom(all_records):
    """Bouwt een hiërarchische themastructuur op basis van domein en onderwerp."""
    tree = {}
    for r in all_records:
        domein = r.get('domein') or 'Niet geclassificeerd'
        onderwerp = r.get('onderwerp_begroting') or 'Overig'

        if domein not in tree:
            tree[domein] = {}
        if onderwerp not in tree[domein]:
            tree[domein][onderwerp] = 0
        tree[domein][onderwerp] += 1

    result = []
    for domein in sorted(tree.keys()):
        children = []
        for onderwerp in sorted(tree[domein].keys()):
            children.append({
                'naam': onderwerp,
                'aantal': tree[domein][onderwerp]
            })
        result.append({
            'naam': domein,
            'kinderen': children,
            'aantal': sum(c['aantal'] for c in children)
        })
    return result


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)

    enriched = read_enriched_sheet(wb)
    print(f"Totaal geschoond: {len(enriched)} records gelezen")

    with open(os.path.join(OUTPUT_DIR, 'raadsbesluiten_verrijkt.json'), 'w', encoding='utf-8') as f:
        json.dump(enriched, f, ensure_ascii=False, indent=2)
    print("  -> raadsbesluiten_verrijkt.json geschreven")

    for year in range(2019, 2027):
        records = read_year_sheet(wb, year)
        if records:
            records = enrich_year_data(records, enriched)
            filename = f'raadsbesluiten_{year}.json'
            with open(os.path.join(OUTPUT_DIR, filename), 'w', encoding='utf-8') as f:
                json.dump(records, f, ensure_ascii=False, indent=2)
            print(f"  -> {filename}: {len(records)} records")

    thema_boom = build_thema_boom(enriched)
    with open(os.path.join(OUTPUT_DIR, 'thema_boom.json'), 'w', encoding='utf-8') as f:
        json.dump(thema_boom, f, ensure_ascii=False, indent=2)
    print(f"  -> thema_boom.json: {len(thema_boom)} domeinen")

    wb.close()


if __name__ == '__main__':
    main()
